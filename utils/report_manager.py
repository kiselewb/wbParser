import json
from typing import AsyncGenerator

import aiofiles
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from config.paths import REPORT_FILE, PRODUCTS_FILE
from config.settings import report_settings


class ReportManager:
    def __init__(self, report_name: str):
        self.report_name = report_name

    async def create_report(self):
        await self._create_main_report()
        await self._create_part_report()

    async def _create_main_report(self):
        wb = Workbook()
        ws = wb.active
        ws.title = self.report_name

        logger.info(f"Создание отчета {ws.title}")

        self._set_headers(ws, report_settings.HEADERS)
        await self._write_data(ws)
        self._set_column_format(ws)

        wb.save(f"{REPORT_FILE}/{self.report_name}.xlsx")

        logger.info(f"Отчет {ws.title} сохранен")

    async def _create_part_report(self):
        wb = Workbook()
        ws = wb.active
        ws.title = f"{self.report_name}_part"

        logger.info(f"Создание отчета {ws.title}")

        self._set_headers(ws, report_settings.HEADERS)
        await self._write_filtered_data(ws)
        self._set_column_format(ws)

        wb.save(f"{REPORT_FILE}/{self.report_name}_part.xlsx")

        logger.info(f"Отчет {ws.title} сохранен")

    @staticmethod
    def _set_headers(ws: Worksheet, headers: list[str]):
        logger.info("Установка и форматирование заголовков")
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            ws.column_dimensions[cell.column_letter].width = (
                len(str(cell.value or "")) + 10
            )

    async def _write_data(self, ws: Worksheet):
        logger.info("Запись данных...")
        async for data in self._data_generator():
            ws.append(self._row_from_data(data))

    async def _write_filtered_data(self, ws: Worksheet):
        logger.info("Запись данных...")
        async for data in self._data_generator():
            is_russian_manufactured = any(
                option["name"] == "Страна производства"
                and option.get("value") == "Россия"
                for option in data["options"]
            )

            if (
                data["rating"] >= 4.5
                and data["price"] <= 10000
                and is_russian_manufactured
            ):
                ws.append(self._row_from_data(data))

    @staticmethod
    def _row_from_data(data: dict) -> list:
        options = [f"{option['name']}: {option['value']}" for option in data["options"]]
        return [
            data["link"],
            data["product_id"],
            data["title"],
            data["price"],
            data["description"],
            ", ".join(data["images"]),
            ", ".join(options),
            data["seller_name"],
            data["seller_link"],
            data["sizes"],
            data["quantity"],
            data["rating"],
            data["reviews_count"],
        ]

    @staticmethod
    def _set_column_format(ws: Worksheet):
        for cell in ws["D"][1:]:
            cell.number_format = "#,##0.00"

    @staticmethod
    async def _data_generator() -> AsyncGenerator[dict]:
        async with aiofiles.open(PRODUCTS_FILE, encoding="utf-8") as f:
            async for line in f:
                yield json.loads(line)
